# =====================================
# Gradient Boosting — randomized search + non-crucial info (FAST PROFILE)
# =====================================
# Unchanged outputs:
#   GradientBoosting.xlsx (sheet "Search")
#   gb_best_params.csv, gb_best_cv.csv, gb_confusion.csv, gb_predictions_fold5.csv
#   gb_top5_truthful.csv, gb_top5_deceptive.csv, gradBoost-accuracies.csv
#   non_crucial_* CSV files
#
# NEW / MODIFIED (speed-ups, duplicates removed):
#   gradBoost-accuracies-v2.csv          ← many rows (fold1..fold5 + test_accuracy)
#   gradBoost-accuracies-v2-params.csv   ← same + GB hyperparameters + vectorizer info
#   hard_replace_search_sheet()          ← keeps Excel "Search" clean every run (single write)
#   RS CV uses shuffle=True, random_state=RNG_SEED (matches original behavior)
#   No separate top-5 retraining; reuse v2 fits to build gb_fold5_validation.csv
#   v2 re-fits are parallelized with joblib
# =====================================

import os, re, string, warnings
from datetime import datetime
from typing import Dict, Tuple, List

import numpy as np
import pandas as pd

from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize

from sklearn.model_selection import StratifiedKFold, RandomizedSearchCV
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, confusion_matrix
from sklearn.ensemble import GradientBoostingClassifier
from scipy.stats import loguniform, randint, uniform

from joblib import Parallel, delayed  # parallel speed-up

warnings.filterwarnings("ignore", category=UserWarning)

# -----------------------------
# Paths
# -----------------------------
DATA_ROOT = r"C:\Users\Thijm\Downloads\op_spam_v1.4 (1)"
OP_ROOT   = os.path.join(DATA_ROOT, "op_spam_v1.4")
CSV_PATH  = "dataset_df.csv"
EXCEL_PATH = "GradientBoosting.xlsx"

if not os.path.isdir(OP_ROOT):
    raise RuntimeError(f"'op_spam_v1.4' not found at: {OP_ROOT}\nFix DATA_ROOT if needed.")
if not os.path.exists(CSV_PATH):
    raise FileNotFoundError("dataset_df.csv must be next to this script/notebook.")

# -----------------------------
# Config  (FAST PROFILE)
# -----------------------------
RNG_SEED     = 42
N_CV_SEARCH  = 3         # CV used in the vectorizer/feature search table
N_CV_FINAL   = 5         # CV used inside RandomizedSearchCV  ← k will equal this

# fast profile changes ↓↓↓
N_ITER_RS = 60                 # was 100
TOP_K_V2_PER_COMBO = 20        # was 50
FEATURE_SIZES = [500, 1000]    # was [200, 500, 1000]
RUN_ALL_COMBOS_FOR_V2 = True   # keep True to cover all combos

VECTORIZERS   = ["count", "tfidf"]
NGRAMS        = [(1,1), (1,2)]
MIN_DF        = 0.02

# Small GB used during the vectorizer search
BASE_GB = dict(n_estimators=150, learning_rate=0.08, max_depth=3, subsample=0.9, min_samples_leaf=2)

# Distributions for RandomizedSearchCV
GB_PARAM_DIST = {
    "loss":               ["log_loss"],
    "n_estimators":       randint(80, 280),
    "learning_rate":      loguniform(0.03, 0.2),
    "max_depth":          randint(2, 5),      # {2,3,4}
    "subsample":          uniform(0.75, 0.25),
    "min_samples_leaf":   randint(1, 6),
    "min_samples_split":  randint(2, 12),
    "max_features":       [None, "sqrt", "log2", 0.5],
    "n_iter_no_change":   [None, 5],
    "validation_fraction":[0.1],
}

# -----------------------------
# Preprocessing
# -----------------------------
try:
    _ = stopwords.words("english")
    _ = word_tokenize("test")
except LookupError:
    import nltk
    nltk.download("stopwords")
    nltk.download("punkt")

stop_words = set(stopwords.words("english"))
stop_words.difference_update({"i","my","we","us","not","never","is","are","was","were","could","would","might","should","very"})
PUNCT_TABLE = str.maketrans("", "", string.punctuation)

def preprocess_text(raw: str) -> str:
    s = (raw or "").translate(PUNCT_TABLE).lower()
    s = re.sub(r"\d+", "", s)
    toks = word_tokenize(re.sub(r"\s+", " ", s).strip())
    toks = [t for t in toks if t not in stop_words]
    return " ".join(toks)

# -----------------------------
# Load data (negative only) + split
# -----------------------------
def load_data() -> Tuple[pd.DataFrame, List[str], np.ndarray, List[str], np.ndarray, List[str]]:
    df = pd.read_csv(CSV_PATH)
    for col in ("fold","txt_path","polarity"):
        if col not in df.columns: raise ValueError(f"CSV missing column: {col}")
    df = df[df["polarity"].astype(str).str.contains("negative", case=False)].copy()
    if df.empty: raise RuntimeError("No rows after filtering to negative polarity.")

    texts, missing = [], []
    for rel in df["txt_path"].astype(str):
        ap = rel if os.path.isabs(rel) else os.path.join(DATA_ROOT, rel.replace("/", os.sep))
        if os.path.exists(ap):
            with open(ap, "r", encoding="utf-8", errors="ignore") as f:
                texts.append(f.read())
        else:
            texts.append(""); missing.append(ap)
    if missing: raise RuntimeError(f"Could not read {len(missing)} files. First missing:\n{missing[0]}")

    df["text"] = texts
    df["text_processed"] = df["text"].map(preprocess_text)
    df["fold"] = df["fold"].astype(str)
    df["label"] = df["txt_path"].str.lower().map(lambda p: 0 if "deceptive" in p else 1)

    is_fold5     = df["fold"].str.lower().eq("fold5")
    X_train_text = df.loc[~is_fold5, "text_processed"].tolist()
    y_train      = df.loc[~is_fold5, "label"].to_numpy()
    X_test_text  = df.loc[ is_fold5, "text_processed"].tolist()
    y_test       = df.loc[ is_fold5, "label"].to_numpy()
    test_keys    = df.loc[ is_fold5, "txt_path"].tolist()
    if len(X_test_text) == 0: raise RuntimeError("No rows in fold5 test split.")

    return df, X_train_text, y_train, X_test_text, y_test, test_keys

# -----------------------------
# Vectorizers
# -----------------------------
def build_vectorizer(kind: str, ngram_range=(1,1), max_features=1000, min_df=MIN_DF):
    common = dict(preprocessor=None, lowercase=False,
                  token_pattern=r"(?u)\b\w+\b", strip_accents="unicode",
                  ngram_range=ngram_range, min_df=min_df, max_features=max_features)
    if kind == "count": return CountVectorizer(**common)
    if kind == "tfidf": return TfidfVectorizer(use_idf=True, norm="l2", **common)
    raise ValueError(f"Unknown vectorizer: {kind}")

# -----------------------------
# 1) Vectorizer/feature/n-gram search (folds 1–4) → Excel (single write)
# -----------------------------
def vectorizer_search(X_texts, y) -> pd.DataFrame:
    """Return df; writing handled by hard_replace_search_sheet to avoid duplicates."""
    cv = StratifiedKFold(n_splits=N_CV_SEARCH, shuffle=True, random_state=RNG_SEED)
    rows = []
    for vec_kind in VECTORIZERS:
        for max_feats in FEATURE_SIZES:
            for ngram in NGRAMS:
                accs = []
                for tr_idx, va_idx in cv.split(X_texts, y):
                    vec = build_vectorizer(vec_kind, ngram_range=ngram, max_features=max_feats)
                    Xtr = vec.fit_transform([X_texts[i] for i in tr_idx]).toarray()
                    Xva = vec.transform([X_texts[i] for i in va_idx]).toarray()
                    clf = GradientBoostingClassifier(random_state=RNG_SEED, **BASE_GB)
                    clf.fit(Xtr, y[tr_idx])
                    accs.append(accuracy_score(y[va_idx], clf.predict(Xva)))
                rows.append({
                    "Vectorizer": vec_kind,
                    "NumberOfFeatures": max_feats,
                    "Unigrams/Bigrams": "uni" if ngram==(1,1) else "uni+bi",
                    **{f"val_acc_fold{i+1}": accs[i] for i in range(len(accs))},
                    "mean_val_acc": float(np.mean(accs)),
                    "std_val_acc":  float(np.std(accs, ddof=1)) if len(accs)>1 else 0.0
                })
    return pd.DataFrame(rows)

def hard_replace_search_sheet(excel_path: str, df_search: pd.DataFrame):
    """Force 'Search' sheet to contain ONLY df_search (prevents mixed/tail rows)."""
    import openpyxl
    from openpyxl import Workbook
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        if "Search" in wb.sheetnames:
            wb.remove(wb["Search"])
        ws = wb.create_sheet("Search")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Search"
    ws.append(list(df_search.columns))
    for _, row in df_search.iterrows():
        ws.append(list(row.values))
    wb.save(excel_path)

def pick_best_by_group(df_search: pd.DataFrame) -> Dict[str, Dict]:
    df = df_search.copy()
    df["_tie_feats"] = df["NumberOfFeatures"]; df["_tie_uni"] = (df["Unigrams/Bigrams"]=="uni").astype(int)
    best_uni = (df[df["Unigrams/Bigrams"]=="uni"]
                .sort_values(["mean_val_acc","_tie_feats","_tie_uni"], ascending=[False,False,False]).iloc[0])
    best_unibi = (df[df["Unigrams/Bigrams"]=="uni+bi"]
                  .sort_values(["mean_val_acc","_tie_feats","_tie_uni"], ascending=[False,False,False]).iloc[0])
    to_combo = lambda r: dict(vec_kind=str(r["Vectorizer"]),
                              max_feats=int(r["NumberOfFeatures"]),
                              ngram=(1,1) if r["Unigrams/Bigrams"]=="uni" else (1,2))
    return {"uni": to_combo(best_uni), "unibi": to_combo(best_unibi)}

# -----------------------------
# 2) RandomizedSearchCV for a given combo
# -----------------------------
def randomized_for_combo(X_texts, y, X_test_texts, y_test, combo) -> Dict:
    vec = build_vectorizer(combo["vec_kind"], ngram_range=combo["ngram"], max_features=combo["max_feats"])
    Xtr_sp = vec.fit_transform(X_texts); Xte_sp = vec.transform(X_test_texts)
    Xtr, Xte = Xtr_sp.toarray(), Xte_sp.toarray()

    rs = RandomizedSearchCV(
        estimator=GradientBoostingClassifier(random_state=RNG_SEED),
        param_distributions=GB_PARAM_DIST,
        n_iter=N_ITER_RS,
        scoring="accuracy",
        cv=StratifiedKFold(n_splits=N_CV_FINAL, shuffle=True, random_state=RNG_SEED),  # match original behavior
        n_jobs=-1, refit=True, random_state=RNG_SEED, verbose=0
    )
    rs.fit(Xtr, y)

    res = pd.DataFrame(rs.cv_results_).sort_values("mean_test_score", ascending=False).copy()
    res["ngrams"] = "uni" if combo["ngram"]==(1,1) else "uni+bi"
    res["Vectorizer"] = combo["vec_kind"]
    res["NumberOfFeatures"] = combo["max_feats"]

    # CV split scores for the best row — ensure fold1..fold5 order
    best_row = res.iloc[0]
    split_cols = sorted([c for c in res.columns if c.startswith("split") and c.endswith("_test_score")],
                        key=lambda s: int(re.search(r"split(\d+)_", s).group(1)))
    cv_accuracies = [float(best_row[c]) for c in split_cols]

    # Train/test metrics for refit best
    best_est = rs.best_estimator_
    train_acc = accuracy_score(y, best_est.predict(Xtr))
    y_pred = best_est.predict(Xte)
    p_truthful = best_est.predict_proba(Xte)[:, 1] if hasattr(best_est, "predict_proba") else np.full(len(y_pred), np.nan)

    test_acc  = accuracy_score(y_test, y_pred)
    test_prec = precision_score(y_test, y_pred, zero_division=0)
    test_rec  = recall_score(y_test, y_pred, zero_division=0)
    test_f1   = f1_score(y_test, y_pred, zero_division=0)

    return {
        "search": rs, "best_est": best_est, "vec": vec,
        "Xtr": Xtr, "Xte": Xte,
        "y_train": y, "y_test": y_test,
        "res_df": res,                   # full, sorted RS table
        "cv_accuracies": cv_accuracies,
        "mean_cv_acc": float(best_row["mean_test_score"]),
        "std_cv_acc":  float(best_row["std_test_score"]),
        "metrics": dict(train_acc=train_acc, test_acc=test_acc, test_prec=test_prec, test_rec=test_rec, test_f1=test_f1),
        "y_pred": y_pred, "p_truthful": p_truthful,
        "combo": combo
    }

# -----------------------------
# Export winner artifacts (legacy kept)
# -----------------------------
def export_winner(out, y, y_test, test_keys):
    best_est  = out["best_est"]; vec = out["vec"]; Xtr = out["Xtr"]
    cv_accs   = out["cv_accuracies"]; mean_cv = out["mean_cv_acc"]; std_cv = out["std_cv_acc"]
    mets      = out["metrics"]; combo = out["combo"]; y_pred = out["y_pred"]; p_truthful = out["p_truthful"]

    # confusion (fold 5)
    cm = confusion_matrix(y_test, y_pred, labels=[0,1])
    pd.DataFrame(cm, index=["true_deceptive(0)", "true_truthful(1)"],
                 columns=["pred_deceptive(0)", "pred_truthful(1)"]).to_csv("gb_confusion.csv", index=True)

    # predictions (fold 5)
    pd.DataFrame({"txt_path": test_keys, "y_true": y_test, "y_pred": y_pred, "p_truthful": p_truthful}) \
        .to_csv("gb_predictions_fold5.csv", index=False)

    # best CV splits
    pdf = pd.DataFrame({"split": list(range(1, len(cv_accs)+1)), "val_accuracy": cv_accs})
    pdf.loc[len(pdf)] = ["mean", mean_cv]; pdf.loc[len(pdf)] = ["std", std_cv]
    pdf.to_csv("gb_best_cv.csv", index=False)

    # legacy summary (unchanged)
    summary = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "vectorizer": combo["vec_kind"], "num_features": combo["max_feats"],
        "ngrams": "uni" if combo["ngram"]==(1,1) else "uni+bi",
        "mean_val_acc": round(mean_cv,4),
        "train_accuracy": round(mets["train_acc"],4), "test_accuracy": round(mets["test_acc"],4),
        "test_precision": round(mets["test_prec"],4), "test_recall": round(mets["test_rec"],4), "test_f1": round(mets["test_f1"],4),
        **out["search"].best_params_, "fold": "fold5"
    }
    pd.DataFrame([summary]).to_csv(
        "gradBoost-accuracies.csv",
        mode=("a" if os.path.exists("gradBoost-accuracies.csv") else "w"),
        header=(not os.path.exists("gradBoost-accuracies.csv")),
        index=False
    )

# -----------------------------
# v2 CSV writers — EXACT LogRegr schema (+ parameters file)
# -----------------------------
V2_COLS = [
    "test_accuracy","with_bigrams","execution_time","num_features","k",
    "avg_val_accuracy","fold1","fold2","fold3","fold4","fold5"
]
PARAM_COLS = ["loss","n_estimators","learning_rate","max_depth","subsample",
              "min_samples_leaf","min_samples_split","max_features","n_iter_no_change"]
EXTRA_INFO_COLS = ["vectorizer","ngrams","rank_by_cv","cv_mean_accuracy","cv_std_accuracy"]

def _eval_candidate(params, Xtr, y, Xte, y_test):
    mdl = GradientBoostingClassifier(**{**params, "random_state": RNG_SEED})
    mdl.fit(Xtr, y)
    pred = mdl.predict(Xte)
    return float(accuracy_score(y_test, pred))

def make_v2_rows_for_combo(out: Dict, top_k: int, k_value: int) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Return (v2_rows, v2_with_params_rows) for the top_k RS candidates of a combo (parallelized)."""
    res = out["res_df"]
    Xtr, y = out["Xtr"], out["y_train"]
    Xte, y_test = out["Xte"], out["y_test"]
    combo = out["combo"]

    split_cols = sorted(
        [c for c in res.columns if c.startswith("split") and c.endswith("_test_score")],
        key=lambda s: int(re.search(r"split(\d+)_", s).group(1))
    )

    # Take top_k rows
    top = res.head(top_k).copy()
    params_list = [r["params"] for _, r in top.iterrows()]

    # Parallel test accuracies on held-out fold5
    test_accs = Parallel(n_jobs=-1, prefer="processes")(
        delayed(_eval_candidate)(p, Xtr, y, Xte, y_test) for p in params_list
    )

    v2_rows, v2p_rows = [], []
    for rank, ((_, r), test_acc) in enumerate(zip(top.iterrows(), test_accs), start=1):
        params = r["params"].copy()
        cv_accs = [float(r[c]) for c in split_cols]
        cv_accs = [round(a,4) for a in cv_accs]
        while len(cv_accs) < 5: cv_accs.append(np.nan)
        if len(cv_accs) > 5: cv_accs = cv_accs[:5]

        v2_row = {
            "test_accuracy": round(float(test_acc), 4),
            "with_bigrams": (combo["ngram"] == (1, 2)),
            "execution_time": datetime.now().isoformat(timespec="seconds"),
            "num_features": int(combo["max_feats"]),
            "k": int(k_value),
            "avg_val_accuracy": round(float(r["mean_test_score"]), 4),
            "fold1": cv_accs[0], "fold2": cv_accs[1], "fold3": cv_accs[2],
            "fold4": cv_accs[3], "fold5": cv_accs[4],
        }
        v2_rows.append(v2_row)

        p_row = {**v2_row}
        p_row["vectorizer"] = combo["vec_kind"]
        p_row["ngrams"] = "uni+bi" if combo["ngram"]==(1,2) else "uni"
        p_row["rank_by_cv"] = rank
        p_row["cv_mean_accuracy"] = round(float(r["mean_test_score"]), 4)
        p_row["cv_std_accuracy"]  = round(float(r.get("std_test_score", np.nan)), 4)
        for kparam in PARAM_COLS:
            p_row[kparam] = params.get(kparam)
        v2p_rows.append(p_row)

    df_v2  = pd.DataFrame(v2_rows, columns=V2_COLS)
    df_v2p = pd.DataFrame(v2p_rows, columns=V2_COLS + EXTRA_INFO_COLS + PARAM_COLS)
    return df_v2, df_v2p

def append_csv(df: pd.DataFrame, path: str, header_cols: List[str]):
    if os.path.exists(path):
        try:
            prev = pd.read_csv(path)
        except Exception:
            prev = pd.DataFrame(columns=header_cols)
        if list(prev.columns) != header_cols:
            df.to_csv(path, index=False)
        else:
            pd.concat([prev, df], ignore_index=True).to_csv(path, index=False)
    else:
        df.to_csv(path, index=False)

# -----------------------------
# Non-crucial info analysis (fold 5 only)
# -----------------------------
def _split_sentences(text: str) -> List[str]:
    sents = re.split(r'(?<=[.!?])\s+', (text or "").strip())
    return [s for s in sents if s]

def add_non_crucial_analysis(
    df_full: pd.DataFrame,
    ngram_range: Tuple[int, int] = (1, 2),
    topn_sentences: int = 5,
    max_fake_to_show: int = 5,
    out_prefix: str = "non_crucial_data_driven",
    text_column: str = "text"
) -> pd.DataFrame:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity as _cs

    df = df_full.copy()
    need_cols = {"fold", "txt_path", text_column}
    miss = need_cols - set(df.columns)
    if miss:
        raise ValueError(f"Non-crucial analysis needs columns: {miss}")

    df["label"] = df["txt_path"].str.lower().map(lambda p: 0 if "deceptive" in p else 1)
    is_test  = df["fold"].astype(str).str.lower().eq("fold5")
    is_train = ~is_test

    train_texts = df.loc[is_train, text_column].astype(str).tolist()
    truthful_train = df.loc[is_train & (df["label"] == 1), text_column].astype(str).tolist()
    if not train_texts or not truthful_train:
        df_full["non_crucial"] = np.nan
        return df_full

    tfidf = TfidfVectorizer(lowercase=True, strip_accents="unicode", stop_words="english",
                            ngram_range=ngram_range, min_df=1).fit(train_texts)

    centroid = np.asarray(tfidf.transform(truthful_train).mean(axis=0)).ravel()[None, :]

    X_test_only = tfidf.transform(df_full.loc[is_test, text_column].astype(str).tolist())
    sims_test   = _cs(X_test_only, centroid).ravel()
    df_full["non_crucial"] = np.nan
    df_full.loc[is_test, "non_crucial"] = 1.0 - sims_test

    # Summary (fold 5)
    fake_nc = df_full.loc[is_test & (df_full["label"] == 0), "non_crucial"].to_numpy()
    real_nc = df_full.loc[is_test & (df_full["label"] == 1), "non_crucial"].to_numpy()
    pd.DataFrame([
        {"class": "fake(0)",     "n": len(fake_nc), "mean_non_crucial": float(np.mean(fake_nc)) if len(fake_nc) else np.nan},
        {"class": "truthful(1)", "n": len(real_nc), "mean_non_crucial": float(np.mean(real_nc)) if len(real_nc) else np.nan},
    ]).to_csv("non_crucial_fold5_summary.csv", index=False)

    # Top deceptive fold-5 reviews and their least-aligned sentences
    top_fakes = (
        df_full.loc[is_test & (df_full["label"] == 0), ["txt_path", text_column, "non_crucial"]]
          .sort_values("non_crucial", ascending=False)
          .head(max_fake_to_show)
    )
    rows = []
    for _, row in top_fakes.iterrows():
        fname = os.path.basename(row["txt_path"])
        sents = _split_sentences(str(row[text_column]))
        if not sents:
            continue
        Xs = tfidf.transform(sents)
        nc_scores = 1.0 - _cs(Xs, centroid).ravel()
        for rank, (score, sent) in enumerate(sorted(zip(nc_scores, sents), key=lambda t: t[0], reverse=True)[:topn_sentences], start=1):
            rows.append({
                "txt_path": row["txt_path"], "file": fname,
                "sentence_rank": rank, "sentence_score": float(score),
                "sentence": re.sub(r"\s+", " ", sent).strip()
            })
    if rows:
        pd.DataFrame(rows).to_csv("non_crucial_top_sentences_fold5.csv", index=False)

    out_csv = f"{out_prefix}_negative_fold5.csv"
    df_full.loc[is_test, ["fold", "txt_path", "label", "non_crucial"]].to_csv(out_csv, index=False)
    return df_full

# -----------------------------
# Run everything
# -----------------------------
if __name__ == "__main__":
    # Load data
    df, X_train_text, y_train, X_test_text, y_test, test_keys = load_data()

    # 1) Vectorizer/feature search on folds 1–4 → Excel (single write)
    df_search = vectorizer_search(X_train_text, y_train)
    hard_replace_search_sheet(EXCEL_PATH, df_search)

    # Build combos to run
    if RUN_ALL_COMBOS_FOR_V2:
        combos_to_run = [
            {"vec_kind": vk, "max_feats": mf, "ngram": ng}
            for vk in VECTORIZERS
            for mf in FEATURE_SIZES
            for ng in NGRAMS
        ]
    else:
        combos = pick_best_by_group(df_search)  # {'uni': {...}, 'unibi': {...}}
        combos_to_run = [combos["uni"], combos["unibi"]]

    # 2) RandomizedSearchCV for each combo
    outs = []
    for combo in combos_to_run:
        outs.append(randomized_for_combo(X_train_text, y_train, X_test_text, y_test, combo))

    # Save TOP-5 per combo (union) — built from RS table (no extra fits)
    def top5_from_res(o: Dict) -> pd.DataFrame:
        res = o["res_df"].head(5).copy()
        def gp(row):
            return {
                "n_estimators": row.get("param_n_estimators", np.nan),
                "learning_rate": row.get("param_learning_rate", np.nan),
                "max_depth": row.get("param_max_depth", np.nan),
                "subsample": row.get("param_subsample", np.nan),
                "min_samples_leaf": row.get("param_min_samples_leaf", np.nan),
                "min_samples_split": row.get("param_min_samples_split", np.nan),
                "max_features": row.get("param_max_features", np.nan),
                "n_iter_no_change": row.get("param_n_iter_no_change", np.nan),
            }
        rows=[]
        for rank, (_, r) in enumerate(res.iterrows(), start=1):
            rows.append({
                "ngrams": "uni" if o["combo"]["ngram"]==(1,1) else "uni+bi",
                "rank": rank,
                "Vectorizer": o["combo"]["vec_kind"],
                "NumberOfFeatures": o["combo"]["max_feats"],
                "mean_val_acc": float(r["mean_test_score"]),
                "std_val_acc": float(r["std_test_score"]),
                **gp(r)
            })
        return pd.DataFrame(rows)
    pd.concat([top5_from_res(o) for o in outs], ignore_index=True).to_csv("gb_best_params.csv", index=False)

    # ===== NEW: write MANY v2 rows (top-K per combo, parallel) =====
    v2_frames, v2p_frames = [], []
    for o in outs:
        v2_df, v2p_df = make_v2_rows_for_combo(o, TOP_K_V2_PER_COMBO, k_value=N_CV_FINAL)
        v2_frames.append(v2_df); v2p_frames.append(v2p_df)
    v2_all  = pd.concat(v2_frames,  ignore_index=True)
    v2p_all = pd.concat(v2p_frames, ignore_index=True)

    append_csv(v2_all,  "gradBoost-accuracies-v2.csv",        V2_COLS)
    append_csv(v2p_all, "gradBoost-accuracies-v2-params.csv", V2_COLS + EXTRA_INFO_COLS + PARAM_COLS)
    # ===============================================================

    # Rebuild gb_fold5_validation.csv and gb_best_validation_params.csv from v2-params (no extra fits)
    fold5_eval_all = (
        v2p_all.sort_values(["vectorizer","ngrams","num_features","rank_by_cv"])
               .groupby(["vectorizer","ngrams","num_features"], as_index=False)
               .head(5)
               .rename(columns={"test_accuracy": "fold5_accuracy"})
               [["ngrams","rank_by_cv","vectorizer","num_features",
                 "cv_mean_accuracy","cv_std_accuracy",
                 "fold5_accuracy","loss","n_estimators","learning_rate","max_depth",
                 "subsample","min_samples_leaf","min_samples_split","max_features","n_iter_no_change"]]
    )
    fold5_eval_all.to_csv("gb_fold5_validation.csv", index=False)
    best_by_fold5 = fold5_eval_all.sort_values("fold5_accuracy", ascending=False).head(1)
    best_by_fold5.to_csv("gb_best_validation_params.csv", index=False)

    # Keep “winner” artifacts as before (winner = best mean CV among combos run)
    winner = max(outs, key=lambda o: o["mean_cv_acc"])
    export_winner(winner, y_train, y_test, test_keys)

    # 3) Non-crucial information analysis (fold 5 only)
    _ = add_non_crucial_analysis(df, ngram_range=(1, 2), topn_sentences=5, max_fake_to_show=5, text_column="text")

    print("Done.")
    print(f"V2 rows written (this run): ~{len(combos_to_run)*TOP_K_V2_PER_COMBO}")
    print("Outputs: GradientBoosting.xlsx, gb_best_params.csv, gb_fold5_validation.csv, gb_best_validation_params.csv,")
    print("         gb_best_cv.csv, gb_confusion.csv, gb_predictions_fold5.csv, gradBoost-accuracies.csv,")
    print("         gradBoost-accuracies-v2.csv, gradBoost-accuracies-v2-params.csv, non_crucial_* CSV files")



